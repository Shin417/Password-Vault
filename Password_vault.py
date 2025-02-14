import openpyxl as xl
import pyinputplus as pyp
import random
import string
import os

#special characters set
special_char_set = set(string.punctuation)

key_check = False
menu_state = 0
try:
    if os.path.exists("Vault.xlsx"):
        wb = xl.load_workbook("Vault.xlsx")
    else:
        wb = xl.Workbook()
except:
    print("File not found")


if not 'Password' in wb.sheetnames:
    password_sheet = wb.active
    password_sheet.title = 'Password'

if not 'Setting' in wb.sheetnames:
    setting_sheet = wb.create_sheet('Setting')
    setting_sheet.title = 'Setting'
wb.save("Vault.xlsx")
ss = wb['Setting']
ps = wb['Password']
master_key = ss['A1'].value
#password_sheet = wb.active
name_pass = dict()

def main_menu():
    select = pyp.inputMenu(['Exit', 'Enter the vault'], numbered=True)
    if select == 'Exit':
        return 0
    elif select == 'Enter the vault':
        return 1

def master_key_is_none():
    if master_key is None:
        return True
    else:
        return False

def master_key_validation(key):
    special_char_check = False
    for char in key:
        if char in special_char_set:
            special_char_check = True
            break
    if len(key) < 5 and special_char_check == False:
        print('Master key must be at least 5 characters and have at least one special character')
        return False
    elif len(key) >= 5 and special_char_check == False:
        print('Master key must have at least one special character')
        return False
    elif len(key) < 5 and special_char_check == True:
        print('Master key must be at least 5 character')
        return False
    else:
        return True

def initiate_master_key():
    global master_key
    key = input("Enter your new master key: ")
    if master_key_validation(key):
        master_key = key
        ss['A1'] = master_key
        wb.save("Vault.xlsx")

def master_key_check():
    global key_check
    global menu_state
    while True:
        key = input("Enter your master key. Enter * to Exit: ")
        if key == master_key:
            key_check = True
            menu_state = 1
            print('You have entered the vault!')
            break
        if key == '*':
            key_check = False
            menu_state = 0
            break
        else:
            print('Master key is invalid')
            key_check = False
            menu_state = 0
            continue

def vault_menu():
    print(r'--------------------------------------------')
    select = pyp.inputMenu(['Exit', 'Search password', 'Store new password', 'View all password', 'Delete password'], numbered=True)
    print(r'--------------------------------------------')
    if select == 'Exit':
        return 0
    elif select == 'Search password':
        return 1
    elif select == 'Store new password':
        return 2
    elif select == 'View all password':
        return 3
    elif select == 'Delete password':
        return 4

def name_validation(name):
    if name == '*':
        print('This character can not be used for password name')
        return False
    elif len(name) > 0:
        return True
    else:
        print('Password name can not be empty')
        return False

def password_validation(password):
    if len(password) > 0:
        return True
    else:
        print('Password can not be empty')
        return False

def search_password():
    while True:
        target = input("Enter your target password name. Enter * to exit: ")
        if target in name_pass.keys():
            print(target + ' : ' + name_pass[target])
            break
        elif target == '*':
            break
        else:
            print('Target password name is not found in the vault.')
            continue

def store(name, password):
    name_pass[name] = password
    print('Your password was successfully stored!')

def generate_password(length):
    password = ""
    char_set = list(string.ascii_uppercase)
    special_set = list(string.punctuation)
    char_set.extend(list(string.ascii_lowercase))
    if length <= 5:
        special_number = 2
    elif length <= 10:
        special_number = 3
    else:
        special_number = 4
    for i in range(length - special_number):
        password += random.choice(char_set)

    chars = []
    for i in range(len(password)):
        chars.append(password[i])
    for i in range(special_number):
        random_index = random.randint(0, length-1)
        chars.insert(random_index, random.choice(special_set))
    return ''.join(chars)

def mode_selection(name):
    mode = pyp.inputMenu(['Exit', 'Enter your password manually', 'Generate random password'], numbered=True)
    #manual enter mode
    if mode == 'Enter your password manually':
        while True:
            password = input("Enter your password: ")
            if password_validation(password):
                store(name, password)
                break
            else:
                continue
    #generate password mode
    elif mode == 'Generate random password':
        length = pyp.inputInt(prompt='Enter your password length: ', min=3, max=15)
        password = generate_password(length)
        store(name, password)

def store_password():
    global menu_state
    while True:
        name = input("Enter your password name. Enter * to exit: ")
        if name == '*':
            break
        elif name_validation(name):
            mode_selection(name)
            break
        else:
            continue

def view_passwords():
    if len(name_pass) == 0:
        print('No record has been found')
    else:
        print(r'--------------------------------------------')
        keys = list(name_pass.keys())
        for i in range(len(name_pass)):
            print(str(keys[i]) + " : " + str(name_pass[keys[i]]))
        print(r'--------------------------------------------')

def delete_password():
    global name_pass
    name_pass_copy = dict(name_pass)
    target = input("Enter your password name to delete. Enter * to exit: ")
    if target in name_pass_copy.keys():
        print(target + ' : ' + name_pass[target])
        conform = pyp.inputYesNo(prompt='Do you want to delete this password? Enter yes or no: ')
        if conform == 'yes':
            del name_pass_copy[target]
            name_pass = name_pass_copy
            print('The password has been deleted!')
    else:
        print('The password name is not found in the vault.')

def set_data_to_sheet():
    keys = list(name_pass.keys())
    values = list(name_pass.values())
    for i in range(len(name_pass)):
        ps.cell(row=i+1, column=1).value = keys[i]
        ps.cell(row=i+1, column=2).value = values[i]
    wb.save("Vault.xlsx")

def get_data_from_sheet():
    height = ps.max_row
    for i in range(height):
        if ps.cell(row=i + 1, column=1).value is not None:
            name_pass[ps.cell(row=i+1, column=1).value] = ps.cell(row=i+1, column=2).value

def clear_sheet():
    height = ps.max_row
    for i in range(height):
        ps.cell(row=i+1, column=1).value = None
        ps.cell(row=i+1, column=2).value = None

#main method -->
#In the main menu
while menu_state == 0:
    main_menu_select = main_menu()
    #exit entire system.
    if main_menu_select == 0:
        menu_state = -1
    #check master key
    elif main_menu_select == 1:
        while master_key_is_none():
            initiate_master_key()
        master_key_check()

    #after master key is checked
    while menu_state == 1 and key_check:
        get_data_from_sheet()
        vault_select = vault_menu()
        if vault_select == 0:
            key_check = False
            menu_state = 0
        elif vault_select == 1:
            search_password()
        elif vault_select == 2:
            store_password()
        elif vault_select == 3:
            view_passwords()
        elif vault_select == 4:
            delete_password()
            clear_sheet()
            set_data_to_sheet()
clear_sheet()
set_data_to_sheet()
exit()
